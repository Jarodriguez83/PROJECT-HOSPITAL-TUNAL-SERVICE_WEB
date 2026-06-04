"""
GENERADOR DE SIMULACIÓN DE HORAS PICO – HOSPITAL DEL TUNAL
CREA UN .xlsx COMPLETO CON:
  HOJA 1 – PACIENTES_SIMULACION   : N° Pacientes con todos sus datos
  HOJA 2 – COLA_PRIORIDAD         : Orden de atención resultante del heap
  HOJA 3 – DASHBOARD_GENERAL      : Estadísticas globales y distribuciones
  HOJA 4 – DASHBOARD_POR_TRIAGE   : Resumen por nivel de triage
  HOJA 5 – DASHBOARD_HORAS_PICO   : Distribución temporal de llegadas
"""

import heapq
import random
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# DATOS DE REFERENCIA 
TIPOS_EMERGENCIA = {
    "PARO CARDIORESPIESTARIO":         (1, 60),
    "POLITRAUMATISMO GRAVE":            (1, 90),
    "DIFICULTAD RESPIRATORIA SEVERA":   (2, 45),
    "ACV / DERRAME CEREBRAL":           (2, 50),
    "FRACTURA CON COMPROMISO VASCULAR": (3, 35),
    "DOLOR ABDOMINAL AGUDO":            (3, 30),
    "FIEBRE ALTA CON CONVULSIÓN":       (4, 25),
    "HERIDA CON SANGRADO MODERADO":     (4, 20),
    "DOLOR LEVE / MALESTAR GENERAL":    (5, 15),
    "CONSULTA MENOR (GRIPE, TOS)":      (5, 10),
}

TRIAGE_LABEL = {
    1: "T1 · ROJO",
    2: "T2 · NARANJA",
    3: "T3 · AMARILLO",
    4: "T4 · VERDE",
    5: "T5 · AZUL",
}

TRIAGE_COLORES_FILL = {
    1: "FFCCCC",  
    2: "FFE5CC",  
    3: "FFEECC",  
    4: "CCFFCC",  
    5: "CCE5FF",  
}

TRIAGE_TEXTO = {
    1: "C0392B", 2: "CA6F1E", 3: "B7950B", 4: "1E8449", 5: "1A5276",
}

EPS_LISTA = [
    "NUEVA EPS", "SURA", "SANITAS", "COMPENSAR", "FAMISANAR",
    "COOMEVA", "CAJACOPI", "MEDIMAS", "SALUD TOTAL", "ALIANSALUD",
]

NOMBRES_M = [
    "Carlos", "Andrés", "Felipe", "Jorge", "Sebastián",
    "Daniel", "Miguel", "Luis", "Juan", "David",
    "Alejandro", "Santiago", "Ricardo", "Camilo", "Nicolás",
]
NOMBRES_F = [
    "Laura", "Sofía", "Valentina", "Natalia", "Camila",
    "Daniela", "Juliana", "Mariana", "Paula", "Alejandra",
    "Isabella", "María", "Ana", "Claudia", "Fernanda",
]
APELLIDOS = [
    "Ramírez", "Gómez", "Martínez", "Torres", "Herrera",
    "Ríos", "Medina", "Vargas", "Pardo", "Ospina",
    "Castro", "Jiménez", "Morales", "Rodríguez", "López",
    "García", "Sánchez", "Díaz", "Reyes", "Cruz",
    "Moreno", "Muñoz", "Alvarado", "Bernal", "Cárdenas",
]

DOCTORES = {
    "MED001": ("Carlos Ramírez",     "Medicina de Emergencias"),
    "MED002": ("Laura Gómez",        "Cirugía General"),
    "MED003": ("Andrés Martínez",    "Cardiología"),
    "MED004": ("Sofía Torres",       "Neurología"),
    "MED005": ("Felipe Herrera",     "Traumatología"),
}
# HORAS PICO 
# MAÑANA: 7 - 10 
# TARDE: 12 - 15
# NOCHE: 18 - 21

HORAS_PICO = [
    (7, 10, 25),    # MAÑANA DE 7 - 10 (25%)
    (10, 12, 10),   # MEDIA MAÑANA DE 10 - 12 (10%)
    (12, 15, 30),   # TARDE DE 12 - 15 (30%)
    (15, 18, 20),   # TARDE NOCHE DE 15 - 18 (20%)
    (18, 21, 15),   # NOCHE DE 18 - 21 (15%)
]

# HELPERS DE GENERACIÓN
def hora_aleatoria():
    """Genera datetime de hoy en una franja de hora pico."""
    r = random.random() * 100
    acum = 0
    for h_ini, h_fin, pct in HORAS_PICO:
        acum += pct
        if r <= acum:
            minuto = random.randint(0, (h_fin - h_ini) * 60 - 1)
            return datetime.now().replace(
                hour=h_ini, minute=0, second=0, microsecond=0
            ) + timedelta(minutes=minuto)
    return datetime.now()

def edad_a_fnac(edad):
    hoy = datetime.now()
    return hoy.replace(year=hoy.year - edad).strftime("%Y-%m-%d")

def grupo_edad(edad):
    if edad <= 1:   return "Neonato"
    if edad <= 5:   return "Infante"
    if edad <= 17:  return "Menor de edad"
    if edad <= 59:  return "Adulto"
    if edad <= 74:  return "Adulto mayor"
    return "Anciano"

def peso_prioridad_edad(edad):
    if edad <= 1:   return 0
    if edad <= 5:   return 1
    if 60 <= edad <= 74: return 2
    if edad >= 75:  return 0
    return 3

def tipo_segun_hora(hora):
    """Más T1/T2 en horas pico nocturnas, más T4/T5 en mañana."""
    h = hora.hour
    if h < 10:
        pesos = [5, 10, 25, 35, 25]
    elif h < 15:
        pesos = [10, 15, 30, 30, 15]
    else:
        pesos = [15, 20, 30, 25, 10]
    nivel = random.choices([1,2,3,4,5], weights=pesos, k=1)[0]
    candidatos = [(t, d) for t,(n,d) in TIPOS_EMERGENCIA.items() if n == nivel]
    return random.choice(candidatos)

def generar_pacientes(n=60):
    pacientes = []
    cedulaset = set()
    for i in range(n):
        sexo = random.choice(["M", "F"])
        nombre = (random.choice(NOMBRES_M) if sexo == "M" else random.choice(NOMBRES_F))
        apellido1 = random.choice(APELLIDOS)
        apellido2 = random.choice(APELLIDOS)
        nombre_completo = f"{nombre} {apellido1} {apellido2}"

        while True:
            cedula = str(random.randint(10_000_000, 99_999_999))
            if cedula not in cedulaset:
                cedulaset.add(cedula)
                break
        # TODO: CATEGORIZACIÓN DE LAS EDADES PARA DAR MÁS PESO A LOS EXTREMOS (NEONATOS, ANCIANOS) EN EL TRIAGE
        edad = random.choices(
            range(0, 91),
            weights=[
                *([5]*2),    # 0-1: NEONATOS
                *([3]*4),    # 2-5: INFANTES
                *([1]*12),   # 6-17: MENORES DE EDAD
                *([2]*42),   # 18-59: ADULTOS
                *([4]*15),   # 60-74: ADULTOS MAYORES
                *([3]*16),   # 75-90: ANCIANOS
            ],
            k=1
        )[0]

        hora_reg = hora_aleatoria()
        tipo_emerg, _ = tipo_segun_hora(hora_reg)
        nivel, tiempo = TIPOS_EMERGENCIA[tipo_emerg]

        # ESPERA ESTIMADA ANTES DE SER ATENDIDO 
        espera_estimada = random.randint(0, 120)
        hora_inicio = hora_reg + timedelta(minutes=espera_estimada)
        hora_fin    = hora_inicio + timedelta(minutes=tiempo)

        doc_id = random.choice(list(DOCTORES.keys()))
        doc_nombre = DOCTORES[doc_id][0]

        pacientes.append({
            "turno":              i + 1,
            "nombre":             nombre_completo,
            "cedula":             cedula,
            "sexo":               sexo,
            "edad":               edad,
            "grupo_edad":         grupo_edad(edad),
            "fecha_nacimiento":   edad_a_fnac(edad),
            "eps":                random.choice(EPS_LISTA),
            "telefono":           f"3{random.randint(10,99)}{random.randint(1_000_000,9_999_999)}",
            "tel_emergencia":     f"3{random.randint(10,99)}{random.randint(1_000_000,9_999_999)}",
            "hora_registro":      hora_reg,
            "tipo_emergencia":    tipo_emerg,
            "nivel_triage":       nivel,
            "triage_label":       TRIAGE_LABEL[nivel],
            "tiempo_atencion":    tiempo,
            "espera_estimada":    espera_estimada,
            "hora_inicio":        hora_inicio,
            "hora_fin":           hora_fin,
            "tiempo_total":       espera_estimada + tiempo,
            "doctor_id":          doc_id,
            "doctor_nombre":      doc_nombre,
            "peso_edad":          peso_prioridad_edad(edad),
        })

    return sorted(pacientes, key=lambda p: (p["hora_registro"], p["turno"]))

def calcular_cola_prioridad(pacientes):
    """Simula la cola min-heap exactamente como en proyecto_salud.py."""
    heap = []
    for i, p in enumerate(pacientes):
        clave = (p["nivel_triage"], p["peso_edad"], p["hora_registro"], i)  # i rompe empates
        heapq.heappush(heap, (clave, p))

    cola = []
    pos = 1
    acum = 0
    while heap:
        clave, p = heapq.heappop(heap)
        cola.append({**p, "posicion_cola": pos, "espera_acumulada": acum})
        acum += p["tiempo_atencion"]
        pos += 1
    return cola

# ESTILOS REUTILIZABLES
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def header_style(ws, row, cols, text, fg="1A3A5C", bg="E8F0FE", size=11):
    ws.merge_cells(start_row=row, start_column=cols[0],
                   end_row=row, end_column=cols[1])
    cell = ws.cell(row=row, column=cols[0], value=text)
    cell.font      = font(bold=True, size=size, color="FFFFFF")
    cell.fill      = fill(fg)
    cell.alignment = center()
    return cell

def col_header(ws, row, col, text, bg="2E5FA3"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = font(bold=True, size=9, color="FFFFFF")
    cell.fill      = fill(bg)
    cell.alignment = center()
    cell.border    = border_thin()
    return cell

def data_cell(ws, row, col, value, bg=None, bold=False, color="222222", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, size=9, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    cell.border = border_thin()
    return cell

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# HOJA 1 — PACIENTES_SIMULACION
def hoja_pacientes(wb, pacientes):
    ws = wb.active
    ws.title = "PACIENTES_SIMULACION"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    header_style(ws, 1, (1, 16),
        "SIMULACIÓN HORAS PICO — HOSPITAL DEL TUNAL · URGENCIAS", size=13)
    header_style(ws, 2, (1, 16),
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  60 pacientes  |  Seed 42 (reproducible)",
        fg="2E5FA3", size=10)

    # ENCABEZADOS DE LAS COLUMNAS
    cols = [
        "N° TURNO","NOMBRE COMPLETO","CÉDULA","SEXO","EDAD","GRUPO EDAD",
        "FECHA NAC.","EPS","TELÉFONO","TEL. EMERGENCIA","HORA REGISTRO",
        "TIPO EMERGENCIA","TRIAGE","T. ATENCIÓN (min)",
        "ESPERA EST. (min)","DOCTOR ASIGNADO",
    ]
    for c, h in enumerate(cols, 1):
        col_header(ws, 3, c, h)

    # DATOS
    for r, p in enumerate(pacientes, 4):
        bg_triage = TRIAGE_COLORES_FILL[p["nivel_triage"]]
        data_cell(ws, r, 1,  p["turno"],           bold=True)
        data_cell(ws, r, 2,  p["nombre"],          wrap=True)
        data_cell(ws, r, 3,  p["cedula"])
        data_cell(ws, r, 4,  p["sexo"])
        data_cell(ws, r, 5,  p["edad"])
        data_cell(ws, r, 6,  p["grupo_edad"])
        data_cell(ws, r, 7,  p["fecha_nacimiento"])
        data_cell(ws, r, 8,  p["eps"],             wrap=True)
        data_cell(ws, r, 9,  p["telefono"])
        data_cell(ws, r, 10, p["tel_emergencia"])
        data_cell(ws, r, 11, p["hora_registro"].strftime("%H:%M:%S"))
        data_cell(ws, r, 12, p["tipo_emergencia"], wrap=True)
        data_cell(ws, r, 13, p["triage_label"],
                  bg=bg_triage, color=TRIAGE_TEXTO[p["nivel_triage"]], bold=True)
        data_cell(ws, r, 14, p["tiempo_atencion"])
        data_cell(ws, r, 15, p["espera_estimada"])
        data_cell(ws, r, 16, p["doctor_nombre"],   wrap=True)

    set_col_widths(ws, [8, 28, 14, 6, 6, 14, 12, 14, 14, 14, 12, 36, 14, 15, 15, 24])
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

# HOJA 2 — COLA_PRIORIDAD
def hoja_cola(wb, cola):
    ws = wb.create_sheet("COLA_PRIORIDAD")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    header_style(ws, 1, (1, 10),
        "COLA DE ATENCIÓN — ORDENADA POR PRIORIDAD (MIN-HEAP)", size=12)
    header_style(ws, 2, (1, 10),
        "Algoritmo: (nivel_triage, peso_edad, hora_registro)  →  menor = atiende primero",
        fg="2E5FA3", size=10)

    cols_h = [
        "POS. COLA","N° TURNO","NOMBRE","TRIAGE","TIPO EMERGENCIA",
        "EDAD","GRUPO EDAD","T. ATENCIÓN (min)","ESPERA ACUM. (min)","DOCTOR",
    ]
    for c, h in enumerate(cols_h, 1):
        col_header(ws, 3, c, h)

    for r, p in enumerate(cola, 4):
        nivel = p["nivel_triage"]
        bg    = TRIAGE_COLORES_FILL[nivel]
        txt   = TRIAGE_TEXTO[nivel]
        pos_bg = "FDEBD0" if r % 2 == 0 else "FFFFFF"
        data_cell(ws, r, 1,  p["posicion_cola"],  bold=True, bg=pos_bg)
        data_cell(ws, r, 2,  p["turno"])
        data_cell(ws, r, 3,  p["nombre"],          wrap=True)
        data_cell(ws, r, 4,  p["triage_label"],    bg=bg, color=txt, bold=True)
        data_cell(ws, r, 5,  p["tipo_emergencia"], wrap=True)
        data_cell(ws, r, 6,  p["edad"])
        data_cell(ws, r, 7,  p["grupo_edad"])
        data_cell(ws, r, 8,  p["tiempo_atencion"])
        data_cell(ws, r, 9,  p["espera_acumulada"])
        data_cell(ws, r, 10, p["doctor_nombre"],   wrap=True)

    set_col_widths(ws, [10, 9, 26, 14, 36, 6, 14, 15, 16, 24])

# HOJA 3 — DASHBOARD GENERAL
def hoja_dashboard_general(wb, pacientes, cola):
    ws = wb.create_sheet("DASHBOARD_GENERAL")
    ws.sheet_view.showGridLines = False

    header_style(ws, 1, (1, 10),
        "DASHBOARD GENERAL — SIMULACIÓN HORAS PICO URGENCIAS", size=13)

    row = 3
    header_style(ws, row, (1, 5), "KPIs GLOBALES", fg="1A3A5C")
    header_style(ws, row, (6, 10), "DISTRIBUCIÓN POR TRIAGE", fg="1A3A5C")
    row += 1

    kpis = [
        ("Total pacientes",           len(pacientes)),
        ("T. espera prom. (min)",      round(sum(p["espera_estimada"] for p in pacientes)/len(pacientes), 1)),
        ("T. atención prom. (min)",    round(sum(p["tiempo_atencion"] for p in pacientes)/len(pacientes), 1)),
        ("T. total prom. (min)",       round(sum(p["tiempo_total"] for p in pacientes)/len(pacientes), 1)),
        ("T. espera máx (min)",        max(p["espera_estimada"] for p in pacientes)),
        ("T. total máx (min)",         max(p["tiempo_total"] for p in pacientes)),
        ("Pacientes T1 críticos",      sum(1 for p in pacientes if p["nivel_triage"]==1)),
        ("Doctores en turno",          len(DOCTORES)),
    ]

    for i, (lbl, val) in enumerate(kpis):
        r = row + i
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font = font(bold=True, size=10)
        c1.fill = fill("EBF5FB")
        c1.border = border_thin()
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = font(bold=True, size=12, color="1A3A5C")
        c2.fill = fill("D6EAF8")
        c2.border = border_thin()
        c2.alignment = center()

    # TABLA DE DISTRIBUCIÓN POR TRIAGE
    col_header(ws, row, 6, "NIVEL TRIAGE")
    col_header(ws, row, 7, "N° PACIENTES")
    col_header(ws, row, 8, "% DEL TOTAL")
    col_header(ws, row, 9, "T. ATENC. PROM.")
    col_header(ws, row, 10, "T. ESPERA PROM.")

    niveles = [1, 2, 3, 4, 5]
    for i, n in enumerate(niveles):
        grupo = [p for p in pacientes if p["nivel_triage"] == n]
        r_n   = row + i + 1
        bg    = TRIAGE_COLORES_FILL[n]
        txt   = TRIAGE_TEXTO[n]
        data_cell(ws, r_n, 6,  TRIAGE_LABEL[n],            bg=bg, color=txt, bold=True)
        data_cell(ws, r_n, 7,  len(grupo))
        pct = round(len(grupo)/len(pacientes)*100, 1) if pacientes else 0
        data_cell(ws, r_n, 8,  f"{pct}%")
        tp  = round(sum(p["tiempo_atencion"] for p in grupo)/len(grupo), 1) if grupo else 0
        te  = round(sum(p["espera_estimada"] for p in grupo)/len(grupo), 1) if grupo else 0
        data_cell(ws, r_n, 9,  f"{tp} min")
        data_cell(ws, r_n, 10, f"{te} min")

    # TABLA DE DOCTORES
    row2 = row + 8
    header_style(ws, row2, (1, 5), "CARGA POR DOCTOR", fg="1A3A5C")
    row2 += 1
    col_header(ws, row2, 1, "ID")
    col_header(ws, row2, 2, "NOMBRE")
    col_header(ws, row2, 3, "ESPECIALIDAD")
    col_header(ws, row2, 4, "N° PACIENTES")
    col_header(ws, row2, 5, "T. TOTAL CARGA (min)")

    for i, (did, (dnombre, desp)) in enumerate(DOCTORES.items()):
        grupo = [p for p in pacientes if p["doctor_id"] == did]
        r_d   = row2 + i + 1
        alt   = "EBF5FB" if i % 2 == 0 else "FFFFFF"
        data_cell(ws, r_d, 1, did,      bg=alt)
        data_cell(ws, r_d, 2, dnombre,  bg=alt, wrap=True)
        data_cell(ws, r_d, 3, desp,     bg=alt, wrap=True)
        data_cell(ws, r_d, 4, len(grupo), bg=alt, bold=True)
        carga = sum(p["tiempo_atencion"] for p in grupo)
        data_cell(ws, r_d, 5, carga,    bg=alt)

    set_col_widths(ws, [28, 14, 14, 14, 18, 14, 22, 14, 14, 18])

    # GRÁFICO DE DISTRIBUCIÓN POR TRIAGE
    conteos_triage = [sum(1 for p in pacientes if p["nivel_triage"]==n) for n in [1,2,3,4,5]]
    aux_start = row + 7 + len(DOCTORES) + 3
    ws.cell(row=aux_start, column=1, value="NIVEL").font = font(bold=True)
    ws.cell(row=aux_start, column=2, value="CANT").font  = font(bold=True)
    for i, (n, c) in enumerate(zip([1,2,3,4,5], conteos_triage)):
        ws.cell(row=aux_start+1+i, column=1, value=TRIAGE_LABEL[n])
        ws.cell(row=aux_start+1+i, column=2, value=c)

    pie = PieChart()
    pie.title = "Distribución por Nivel de Triage"
    pie.style = 10
    labels    = Reference(ws, min_col=1, min_row=aux_start+1, max_row=aux_start+5)
    data_ref  = Reference(ws, min_col=2, min_row=aux_start,   max_row=aux_start+5)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.width  = 16
    pie.height = 12
    ws.add_chart(pie, "E18")

    # GRÁFICO DE BARRAS: CARGA POR DOCTOR
    bar_start = aux_start + 8
    ws.cell(row=bar_start, column=1, value="DOCTOR").font  = font(bold=True)
    ws.cell(row=bar_start, column=2, value="PACIENTES").font = font(bold=True)
    for i, (did, (dnombre, _)) in enumerate(DOCTORES.items()):
        grupo = [p for p in pacientes if p["doctor_id"] == did]
        ws.cell(row=bar_start+1+i, column=1, value=dnombre.split()[0])
        ws.cell(row=bar_start+1+i, column=2, value=len(grupo))

    bar = BarChart()
    bar.type  = "col"
    bar.style = 10
    bar.title = "PACIENTES ASIGNADOS POR DOCTOR"
    bar.y_axis.title = "N° Pacientes"
    bar.x_axis.title = "Doctor"
    cats    = Reference(ws, min_col=1, min_row=bar_start+1, max_row=bar_start+len(DOCTORES))
    data_b  = Reference(ws, min_col=2, min_row=bar_start,   max_row=bar_start+len(DOCTORES))
    bar.add_data(data_b, titles_from_data=True)
    bar.set_categories(cats)
    bar.width  = 16
    bar.height = 12
    ws.add_chart(bar, "E30")


# HOJA 4 — DASHBOARD POR PACIENTE
def hoja_dashboard_paciente(wb, cola):
    ws = wb.create_sheet("DASHBOARD_POR_PACIENTE")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    header_style(ws, 1, (1, 12),
        "DASHBOARD POR PACIENTE — TIEMPOS Y PRIORIDAD EN COLA", size=12)
    header_style(ws, 2, (1, 12),
        "Cada fila = un paciente con sus métricas de atención. Coloreado por nivel de triage.",
        fg="2E5FA3", size=10)

    cols_h = [
        "POS. COLA","N° TURNO","NOMBRE","TRIAGE","TIPO EMERGENCIA",
        "EDAD","GRUPO EDAD","HORA REG.","T. ESPERA (min)",
        "T. ATENCIÓN (min)","T. TOTAL (min)","DOCTOR",
    ]
    for c, h in enumerate(cols_h, 1):
        col_header(ws, 3, c, h)

    for r, p in enumerate(cola, 4):
        nivel = p["nivel_triage"]
        bg    = TRIAGE_COLORES_FILL[nivel]
        txt   = TRIAGE_TEXTO[nivel]
        alt   = "F8F9FA" if r % 2 == 0 else "FFFFFF"

        data_cell(ws, r, 1,  p["posicion_cola"],      bold=True, bg=bg, color=txt)
        data_cell(ws, r, 2,  p["turno"],              bg=alt)
        data_cell(ws, r, 3,  p["nombre"],             bg=alt, wrap=True)
        data_cell(ws, r, 4,  p["triage_label"],       bg=bg, color=txt, bold=True)
        data_cell(ws, r, 5,  p["tipo_emergencia"],    bg=alt, wrap=True)
        data_cell(ws, r, 6,  p["edad"],               bg=alt)
        data_cell(ws, r, 7,  p["grupo_edad"],         bg=alt)
        data_cell(ws, r, 8,  p["hora_registro"].strftime("%H:%M:%S"), bg=alt)
        data_cell(ws, r, 9,  p["espera_acumulada"],   bg=alt)
        data_cell(ws, r, 10, p["tiempo_atencion"],    bg=alt)
        total = p["espera_acumulada"] + p["tiempo_atencion"]
        t_bg = "FFD5D5" if total > 100 else ("FFFACC" if total > 50 else "D5FFD5")
        data_cell(ws, r, 11, total,                   bg=t_bg, bold=True)
        data_cell(ws, r, 12, p["doctor_nombre"],      bg=alt, wrap=True)

    set_col_widths(ws, [10, 9, 26, 14, 34, 6, 14, 10, 14, 15, 12, 24])

# HOJA 5 — DASHBOARD HORAS PICO
def hoja_horas_pico(wb, pacientes):
    ws = wb.create_sheet("DASHBOARD_HORAS_PICO")
    ws.sheet_view.showGridLines = False

    header_style(ws, 1, (1, 8),
        "ANÁLISIS DE HORAS PICO — DISTRIBUCIÓN TEMPORAL DE LLEGADAS", size=12)

    franjas = {}
    for h in range(6, 23):
        franjas[h] = {"total": 0, "t1": 0, "t2": 0, "t3": 0, "t4": 0, "t5": 0}

    for p in pacientes:
        h = p["hora_registro"].hour
        if h in franjas:
            franjas[h]["total"] += 1
            franjas[h][f"t{p['nivel_triage']}"] += 1

    row = 3
    header_style(ws, row, (1, 8), "LLEGADAS POR HORA DEL DÍA", fg="1A3A5C")
    row += 1
    cols_h = ["HORA","TOTAL","T1 ROJO","T2 NARANJA","T3 AMARILLO","T4 VERDE","T5 AZUL","% DEL DÍA"]
    for c, h in enumerate(cols_h, 1):
        col_header(ws, row, c, h)
    row += 1

    total_pac = len(pacientes)
    for h, datos in franjas.items():
        pct = round(datos["total"] / total_pac * 100, 1) if total_pac else 0
        es_pico = any(h_ini <= h < h_fin for h_ini, h_fin, _ in HORAS_PICO)
        bg_row = "FFF3E0" if es_pico else "FFFFFF"
        ws.cell(row=row, column=1, value=f"{h:02d}:00 - {h+1:02d}:00").fill = fill(bg_row)
        ws.cell(row=row, column=1).font = font(bold=es_pico, size=9)
        ws.cell(row=row, column=1).border = border_thin()
        ws.cell(row=row, column=1).alignment = center()

        for c, key in enumerate(["total","t1","t2","t3","t4","t5"], 2):
            v = datos[key]
            nivel_idx = c - 1  # c=3→t1(nivel 1), c=4→t2... but c=2 is total
            nivel_real = c - 1 if c >= 3 else None
            t_bg = (TRIAGE_COLORES_FILL[nivel_real] if nivel_real and nivel_real in TRIAGE_COLORES_FILL and v > 0 else
                    ("E8F8F5" if c == 2 and v > 0 else "F8F9FA"))
            data_cell(ws, row, c, v, bg=t_bg if v > 0 else bg_row)
        data_cell(ws, row, 8, f"{pct}%",
                  bg="FFE0B2" if es_pico else "FFFFFF",
                  bold=es_pico)
        row += 1

    data_cell(ws, row, 1, "TOTAL", bold=True, bg="1A3A5C", color="FFFFFF")
    data_cell(ws, row, 2, f"=SUM(B5:B{row-1})", bold=True, bg="D6EAF8")
    for c in range(3, 8):
        data_cell(ws, row, c, f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{row-1})",
                  bold=True, bg="D6EAF8")
    data_cell(ws, row, 8, "100%", bold=True, bg="D6EAF8")

    row += 2
    # HORAS PICO
    header_style(ws, row, (1, 4), "LEYENDA: HORAS PICO", fg="E67E22")
    row += 1
    picos = [
        ("07:00 - 10:00", "MAÑANA (Alta)", "Llegada de pacientes post-madrugada"),
        ("12:00 - 15:00", "TARDE (Muy Alta)", "Hora de mayor afluencia del día"),
        ("18:00 - 21:00", "NOCHE (Media-Alta)", "Pacientes post-jornada laboral"),
    ]
    for lbl, tipo, desc in picos:
        data_cell(ws, row, 1, lbl,  bold=True, bg="FFF3E0")
        data_cell(ws, row, 2, tipo, bold=True, bg="FDEBD0", color="C0392B")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=3, value=desc)
        cell.font = font(size=9, italic=True)
        cell.border = border_thin()
        row += 1

    # GRÁFICO DE BARRAS POR HORA
    bar_start = 4
    bar = BarChart()
    bar.type  = "col"
    bar.style = 10
    bar.title = "Pacientes por Hora del Día (Simulación Horas Pico)"
    bar.y_axis.title = "N° Pacientes"
    bar.x_axis.title = "Hora"
    cats   = Reference(ws, min_col=1, min_row=bar_start+1, max_row=bar_start+len(franjas))
    data_b = Reference(ws, min_col=2, min_row=bar_start,   max_row=bar_start+len(franjas))
    bar.add_data(data_b, titles_from_data=True)
    bar.set_categories(cats)
    bar.width  = 22
    bar.height = 14
    ws.add_chart(bar, "I3")

    set_col_widths(ws, [22, 8, 10, 12, 13, 11, 10, 10])

# MAIN
def main():
    import sys
    n_pacientes = 60
    seed_valor  = 42

    if len(sys.argv) >= 2:
        try:
            n_pacientes = int(sys.argv[1])
            if n_pacientes < 1:
                raise ValueError
        except ValueError:
            print(" EL PRIMER ARGUMENTO DEBE SER UN NÚMERO ENTERO POSITIVO (N° PACIENTES).")
            print(" EJEMPLO: python generar_simulacion.py 100 42")
            sys.exit(1)

    if len(sys.argv) >= 3:
        try:
            seed_valor = int(sys.argv[2])
            if seed_valor < 0:
                raise ValueError
        except ValueError:
            print(" EL SEGUNDO ARGUMENTO DEBE SER UN NÚMERO ENTERO >= 0 (SEED).")
            print(" EJEMPLO: python generar_simulacion.py 100 42")
            sys.exit(1)

    # SEED 0 → ALEATORIO REAL (Distinto en cada ejecución)
    if seed_valor == 0:
        import time
        seed_valor = int(time.time()) % 100_000
        print(f"SEMILLA ALEATORIA GENERADA AUTOMATICAMENTE: {seed_valor}")

    random.seed(seed_valor)

    sep = "─" * 53
    print(f"\n{sep}")
    print(f"SIMULACIÓN — HOSPITAL DEL TUNAL- SISTEMA SIGU")
    print(f"{sep}")
    print(f"  PACIENTES   : {n_pacientes}")
    print(f"  SEED        : {seed_valor}")
    print(f"{sep}")

    print("GENERANDO LOS DATOS DE SIMULACIÓN...")
    pacientes = generar_pacientes(n_pacientes)
    cola      = calcular_cola_prioridad(pacientes)

    wb = Workbook()
    hoja_pacientes(wb, pacientes)
    hoja_cola(wb, cola)
    hoja_dashboard_general(wb, pacientes, cola)
    hoja_dashboard_paciente(wb, cola)
    hoja_horas_pico(wb, pacientes)

    nombre_archivo = f"SIMUL_PACIENTES{n_pacientes}_SEED{seed_valor}.xlsx"
    import os
    os.makedirs("outputs", exist_ok=True)
    out = f"outputs/{nombre_archivo}"
    wb.save(out)

    # RESUMEN EN CONSOLA 
    conteos     = {n: sum(1 for p in pacientes if p["nivel_triage"] == n) for n in [1,2,3,4,5]}
    prom_espera = round(sum(p["espera_estimada"] for p in pacientes) / len(pacientes), 1)
    prom_total  = round(sum(p["tiempo_total"]    for p in pacientes) / len(pacientes), 1)
    max_espera  = max(p["espera_estimada"] for p in pacientes)
    max_total   = max(p["tiempo_total"]    for p in pacientes)

    print(f"\n  DISTRIBUCIÓN DE TRIAGE")
    for n, c in conteos.items():
        barra = "█" * c + "░" * max(0, 20 - c)
        pct   = round(c / len(pacientes) * 100, 1)
        print(f"  {TRIAGE_LABEL[n]:<18} {barra}  {c:>3} pac  ({pct}%)")

    print(f"\n  MÉTRICAS DE TIEMPO")
    print(f"  ESPERA PROMEDIO: {prom_espera:>6} min")
    print(f"  TIEMPO TOTAL PROMEDIO: {prom_total:>6} min")
    print(f"  ESPERA MÁXIMA: {max_espera:>6} min")
    print(f"  TIEMPO TOTAL MÁXIMO: {max_total:>6} min")

    print(f"\n  COLA DE PRIORIDAD DE TODOS LOS PACIENTES")
    for p in cola[:n_pacientes]: # MOSTRAR LOS N PRIMEROS PACIENTES
        print(f"  #{p['posicion_cola']:>2}  {p['nombre']:<28}  {p['triage_label']}")

    print(f"\n ✔ ARCHIVO GUARDADO: {nombre_archivo}")
    print(f"{sep}\n")

if __name__ == "__main__":
    main()